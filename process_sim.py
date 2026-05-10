"""
Process Simulation – Response Surface Model
============================================
Simuliert prozessnahe Fab-Messdaten mit versteckten Modellparametern.
R_true(x) = R_max - sum_i a_i*(xi-xi*)^2
           + sum_{i<j} b_ij*(xi-xi*)*(xj-xj*)
           + sum_{i<j<l} c_ijl*(xi-xi*)*(xj-xj*)*(xl-xl*)
           + eps_proc
"""

import numpy as np
import json
import os
from itertools import combinations
from openpyxl import load_workbook, Workbook

# ─────────────────────────────────────────────────────────────
# 1) Konfiguration & Konstanten
# ─────────────────────────────────────────────────────────────
A_TOT       = 0.05          # Gesamtstärke der Haupteffekte (fest)
R_MAX       = 100.0         # Maximaler Ergebniswert (fest)
N_FACTORS   = 4             # Anzahl Eingangsvariablen 

SIGMA_BASE  = 2.4           # Erwartetes Prozessrauschen (Mittelwert)
SIGMA_VAR   = 0.3           # ± Streuung um SIGMA_BASE

# Koeffizientenbereiche
COEF_LARGE  = (0.8, 1.5)    # "großer" Koeffizient
COEF_SMALL  = (0.05, 0.3)   # "kleiner" Koeffizient

# Skalierungsgrenzen für Interaktionen
B_SCALE_MAX = 0.4 * A_TOT   # max sum |b_ij|
C_SCALE_MAX = 0.03 * A_TOT   # max sum |c_ijl|

# Optimum-Suchbereich (x* liegt in diesem Bereich)
X_STAR_RANGES = [(10.0, 50.0), (5.0, 25.0), (250.0, 500.0), (70.0, 180.0)]
FACTOR_NAMES = ["Bias_Power_W", "Chamber_Pressure_mTorr", "RF_Power_W", "CHF3_flow_sccm"]
RUN_FILE = "runs.json"      # Datei mit Parametern


# ─────────────────────────────────────────────────────────────
# 2) Zufallshilfen & Basis-Sampling
# ─────────────────────────────────────────────────────────────
def _rand_range(lo, hi):
    return np.random.uniform(lo, hi)

def _rand_signed(lo, hi):
    """Zieht Wert aus [lo, hi] mit zufälligem Vorzeichen."""
    return _rand_range(lo, hi) * np.random.choice([-1, 1])

def _draw_sigma():
    return max(0.1, np.random.normal(SIGMA_BASE, SIGMA_VAR))


# ─────────────────────────────────────────────────────────────
# 3) Erzeugung der Modellkoeffizienten
# ─────────────────────────────────────────────────────────────
def _gen_main_effects():
    """Haupteffekte a_i: 50% groß / 50% klein, Summe = A_TOT."""
    a = []
    for _ in range(N_FACTORS):
        if np.random.rand() < 0.6:
            a.append(_rand_range(*COEF_LARGE))
        else:
            a.append(_rand_range(*COEF_SMALL))
    a = np.array(a)
    a = a / a.sum() * A_TOT          # Normierung: sum(a) = A_TOT

    return np.round(a,3)

def _gen_two_factor(a_sum):
    """b_ij: 30% groß / 70% klein, sum|b| ≤ B_SCALE_MAX."""
    pairs = list(combinations(range(N_FACTORS), 2))
    b = {}
    for p in pairs:
        if np.random.rand() < 0.3:
            b[p] = _rand_signed(*COEF_LARGE)
        else:
            b[p] = _rand_signed(*COEF_SMALL)
    total = sum(abs(v) for v in b.values())
    b_max = 0.3 * a_sum
    if total > b_max:
        scale = b_max / total
        b = {k: v * scale for k, v in b.items()}
    return {k: round(v, 3) for k, v in b.items()}

def _gen_three_factor(a_sum):
    """c_ijl: 15% groß / 85% klein, sum|c| ≤ C_SCALE_MAX."""
    triples = list(combinations(range(N_FACTORS), 3))
    c = {}
    for t in triples:
        if np.random.rand() < 0.15:
            c[t] = _rand_signed(*COEF_LARGE)
        else:
            c[t] = _rand_signed(*COEF_SMALL)
    total = sum(abs(v) for v in c.values())
    c_max = 0.03 * a_sum
    if total > c_max:
        scale = c_max / total
        c = {k: v * scale for k, v in c.items()}
    return {k: round(v, 3) for k, v in c.items()}


# ─────────────────────────────────────────────────────────────
# 4) Erzeugung der unbekannten Optima x*
# ─────────────────────────────────────────────────────────────
def _gen_optima():
    return np.array([round(_rand_range(*r), 1) for r in X_STAR_RANGES])




# ─────────────────────────────────────────────────────────────
# 5) Hinweisbereiche um x* (verschleiert)
# ─────────────────────────────────────────────────────────────
def _gen_hint_ranges(x_star):
    """
    Gibt pro x_i ein Intervall [lo, hi] zurück.
    x* liegt darin, aber NICHT offensichtlich in der Mitte.
    """
    hints = []
    for i, xi in enumerate(x_star):
     if i == 0:
         delta = _rand_range(5, 15)
     elif i == 1:
         delta = _rand_range(3, 10)
     elif i == 2:
         delta = _rand_range(30, 60)
     else:
         delta = _rand_range(15, 40)
     lo = max(round(xi - _rand_range(0,1) * 2 * delta, 1), X_STAR_RANGES[i][0])
     hi = min(round(xi + _rand_range(0,1) * 2 * delta, 1), X_STAR_RANGES[i][1])
     if lo > hi:
         lo, hi = hi, lo
     hints.append((lo, hi))
    return hints
     

# ─────────────────────────────────────────────────────────────
# 6) Berechnung R_true(x) – Modellkern
# ─────────────────────────────────────────────────────────────
def _calc_r(x, a, b, c, x_star, sigma, add_noise=True):
    """
    R_true(x) = R_max
               - sum_i  a_i  * (xi - xi*)^2
               + sum_{i<j} b_ij * (xi-xi*)(xj-xj*)
               + sum_{i<j<l} c_ijl * (xi-xi*)(xj-xj*)(xl-xl*)
               + eps
    """
    x    = np.asarray(x, dtype=float)
    dx   = x - x_star                              # Abweichung vom Optimum

    # Haupteffekte (Penalty)
    main = np.sum(a * dx**2)

    # Zweifaktor-Interaktionen
    two = sum(b[p] * dx[p[0]] * dx[p[1]] for p in b)

    # Dreifaktor-Interaktionen
    three = sum(c[t] * dx[t[0]] * dx[t[1]] * dx[t[2]] for t in c)

    R = R_MAX - main + two + three

    if add_noise:
        R += np.random.normal(0, 0.03*abs(R))

    return float(R)


# ─────────────────────────────────────────────────────────────
# 7) Run-Verwaltung der unbekannten Variablen (Vault) 
# ─────────────────────────────────────────────────────────────
def _serialize_run(run_data):
    """Konvertiert numpy-Typen für JSON."""
    d = {
        "a":      run_data["a"].tolist(),
        "b":      {str(k): float(v) for k, v in run_data["b"].items()},
        "c":      {str(k): float(v) for k, v in run_data["c"].items()},
        "sigma":  float(run_data["sigma"]),
        "x_star": run_data["x_star"].tolist(),
        "hints":  run_data["hints"],
    }
    return d

def _deserialize_run(d):
    import ast
    return {
        "a":      np.array(d["a"]),
        "b":      {ast.literal_eval(k): v for k, v in d["b"].items()},
        "c":      {ast.literal_eval(k): v for k, v in d["c"].items()},
        "sigma":  d["sigma"],
        "x_star": np.array(d["x_star"]),
        "hints":  [tuple(h) for h in d["hints"]],
    }

def _load_all_runs():
    if os.path.exists(RUN_FILE):
        with open(RUN_FILE, "r") as f:
            return json.load(f)
    return {}

def _save_all_runs(runs):
    with open(RUN_FILE, "w") as f:
        json.dump(runs, f, indent=2)

# In-Memory-Vault (aktive Session)
_vault = {}

def create_run(run_id=None, persist=True):
    """
    Erstellt einen neuen Run mit zufälligen Parametern.

    Returns
    -------
    run_id : str
    hints  : list of (lo, hi) pro x_i
    """
    if run_id is None:
        run_id = f"run_{np.random.randint(1000, 9999)}"

    a      = _gen_main_effects()
    a_sum  = a.sum()
    b      = _gen_two_factor(a_sum)
    c      = _gen_three_factor(a_sum)
    sigma  = _draw_sigma()
    x_star = _gen_optima()
    hints  = _gen_hint_ranges(x_star)

    run_data = {"a": a, "b": b, "c": c, "sigma": sigma,
                "x_star": x_star, "hints": hints}
    _vault[run_id] = run_data

    if persist:
        all_runs = _load_all_runs()
        all_runs[run_id] = _serialize_run(run_data)
        _save_all_runs(all_runs)

    print(f"\n{'='*50}")
    print(f"  Run erstellt: {run_id}")
    print(f"{'='*50}")
    print("  Hinweisbereiche für x* (Optimum liegt darin):")
    for i, (lo, hi) in enumerate(hints, 1):
        print(f"    x_{i}: [{lo:.1f}, {hi:.1f}]")
    print(f"{'='*50}\n")

    return run_id, hints


def load_run(run_id):
    """Lädt einen gespeicherten Run aus Datei in den Vault."""
    all_runs = _load_all_runs()
    if run_id not in all_runs:
        raise KeyError(f"Run '{run_id}' nicht gefunden.")
    _vault[run_id] = _deserialize_run(all_runs[run_id])
    print(f"Run '{run_id}' geladen.")
    return _vault[run_id]["hints"]


def reveal_run(run_id):
    """Zeigt die versteckten Modellparameter eines Runs (Vault öffnen)."""
    if run_id not in _vault:
        load_run(run_id)
    rd = _vault[run_id]
    print(f"\n{'='*50}")
    print(f"  VAULT – {run_id}")
    print(f"{'='*50}")
    print(f"  x*     = {np.round(rd['x_star'], 3)}")
    print(f"  a      = {np.round(rd['a'], 4)}")
    print(f"  sigma  = {rd['sigma']:.4f}")
    print("  b_ij:")
    for k, v in rd["b"].items():
        idx = tuple(i+1 for i in k)
        print(f"    b{idx} = {v:.4f}")
    print("  c_ijl:")
    for k, v in rd["c"].items():
        idx = tuple(i+1 for i in k)
        print(f"    c{idx} = {v:.4f}")
    print(f"{'='*50}\n")


# ─────────────────────────────────────────────────────────────
# 8) Messung (einzeln oder Mehrfachmessung)
# ─────────────────────────────────────────────────────────────
def measure(run_id, x, n=1, add_noise=True):
    """
    Erzeugt n Messwerte für festes x.

    Parameters
    ----------
    run_id    : str       – Run-ID
    x         : array-like (4,) – Einstellwerte [x1, x2, x3, x4]
    n         : int       – Anzahl Wiederholungen
    add_noise : bool      – Prozessrauschen an/aus

    Returns
    -------
    results : list of float
    """
    if run_id not in _vault:
        load_run(run_id)
    rd = _vault[run_id]
    x  = np.asarray(x, dtype=float)
    if x.shape != (N_FACTORS,):
        raise ValueError(f"x muss {N_FACTORS} Werte haben.")

    results = [
        _calc_r(x, rd["a"], rd["b"], rd["c"], rd["x_star"],
                rd["sigma"], add_noise=add_noise)
        for _ in range(n)
    ]

    if n == 1:
        print(f"  R = {results[0]:.4f}  (x = {x.tolist()})")
    else:
        arr = np.array(results)
        print(f"  x = {x.tolist()}")
        print(f"  n = {n}  |  mean = {arr.mean():.4f}  |  "
              f"std = {arr.std(ddof=1):.4f}  |  "
              f"range = [{arr.min():.4f}, {arr.max():.4f}]")
    return [round(r,3)for r in results]
        

# ─────────────────────────────────────────────────────────────
# 9) Daten in Excel einfügen lassen
# ─────────────────────────────────────────────────────────────

def export_to_excel(results, filename="Control-Charts.xlsx", start_col=2, row=4, sheet=1):
    """
    Schreibt Messwerte nebeneinander in Excel.
    Standard: B4, C4, D4, ...
    """

    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    ws = wb.worksheets[sheet - 1]

    for i, val in enumerate(results):
        ws.cell(row=row, column=start_col + i, value=val)

    wb.save(filename)
    print(f"  {len(results)} Werte gespeichert in '{filename}' ab Zelle "
          f"{ws.cell(row=row, column=start_col).coordinate}")
# ─────────────────────────────────────────────────────────────
# 10) DoE preparation
# ─────────────────────────────────────────────────────────────

def export_to_csv(results, filename="doe_data.csv"):
    """
    Schreibt Messwerte untereinander in CSV.
    Bestehende Datei wird ergänzt, nicht überschrieben.
    """
    import csv
    file_exists = os.path.exists(filename)
    
    with open(filename, "a", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["R"])          # Header nur beim ersten Mal
        for val in results:
            writer.writerow([val])
    
    print(f"  {len(results)} Werte gespeichert in '{filename}'")
    
    
# ─────────────────────────────────────────────────────────────
# 11) Testen der Streu-Weite der Ober und Untergrenzen
# ─────────────────────────────────────────────────────────────
# run_id, hints = create_run(run_id="Uebung1", persist=True)

# x_nah  = [(lo + hi) / 2 for lo, hi in hints]   # Mitte = nah an x*
# x_lo   = [lo for lo, hi in hints]               # unterer Rand
# x_hi   = [hi for lo, hi in hints]               # oberer Rand

# r_nah = measure(run_id, x_nah, n=1, add_noise=False)
# r_lo  = measure(run_id, x_lo,  n=1, add_noise=False)
# r_hi  = measure(run_id, x_hi,  n=1, add_noise=False)

# snr = (max(r_lo[0], r_hi[0]) - min(r_lo[0], r_hi[0])) / (0.03 * r_nah[0])
# print(f"  Spannweite R: {max(r_lo[0],r_hi[0]):.2f} – {min(r_lo[0],r_hi[0]):.2f}")
# print(f"  SNR: {snr:.1f}  (Ziel: >3)")


# ─────────────────────────────────────────────────────────────
# 12) DoE
# ─────────────────────────────────────────────────────────────
def run_doe(run_id, hints, filename="DoE.csv"):
    """
    Vollfaktorielles 2^4 Design + 4 Center Points.
    +1 = oberer Hinweisbereich, -1 = unterer Hinweisbereich, 0 = Mitte.
    Speichert: x1_coded, x2_coded, x3_coded, x4_coded, R
    """
    import csv
    from itertools import product

    lo = [h[0] for h in hints]
    hi = [h[1] for h in hints]
    mid = [(l + h) / 2 for l, h in zip(lo, hi)]

    def decode(coded):
        return [lo[i] if c == -1 else hi[i] if c == 1 else mid[i]
                for i, c in enumerate(coded)]

# 2^4 = 16 Runs
    runs = list(product([-1, 1], repeat=4))
    # 4 Center Points
    runs += [(0, 0, 0, 0)] * 4
    # Axial Points (Face-Centered: alpha = ±1)
    for i in range(4):
        for sign in [-1, 1]:
            point = [0, 0, 0, 0]
            point[i] = sign
            runs.append(tuple(point))
    with open(filename, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([*FACTOR_NAMES, "R"])
        for coded in runs:
            x_real = decode(coded)
            r = measure(run_id, x_real, n=1, add_noise=True)[0]
            writer.writerow([*coded, r])

    print(f"  DoE abgeschlossen: {len(runs)} Runs gespeichert in '{filename}'")

# ─────────────────────────────────────────────────────────────
# Start
# ─────────────────────────────────────────────────────────────

" run erstellen "
# run_id, hints = create_run(run_id="project_data", persist=True)
" run werte nachsehen "
# hints = load_run("project_data")
# run_id = "project_data"
# run_doe(run_id, hints)

" Variablen festlegen "
# x = [38,14,255,110]

" Ergebnisse bei Variablen "
# measure(run_id,x,n=10)

" Ergebnisse in Excel für SPC "
# x = [25,8,300,125]
# results = measure(run_id,x,n=10)
# export_to_excel(measure(run_id,x,n=10))

" DoE "
# run_doe(run_id, hints)

"neuer Run bei besseren Parametern"
# x = [35.6,22.5,261.3,128.9]
# results = measure(run_id,x,n=15)
# export_to_excel(measure(run_id,x,n=15),sheet=2)

" Lösung einsehen "
# reveal_run(run_id)


